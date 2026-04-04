from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
import csv
from pathlib import Path
from typing import Any

from graph_agent.runtime.run_documents import normalize_run_documents

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency guard
    load_workbook = None


SUPPORTED_SPREADSHEET_FORMATS = {"csv", "xlsx"}
SPREADSHEET_STORAGE_SUFFIXES = {".csv", ".xlsx"}
DEFAULT_SAMPLE_ROW_LIMIT = 5
SPREADSHEET_HEADER_ROW_INDEX = 1
SPREADSHEET_FIRST_DATA_ROW_INDEX = 2


class SpreadsheetParseError(ValueError):
    """Raised when a spreadsheet file cannot be parsed with the current config."""


def resolve_spreadsheet_path_from_run_documents(
    documents: Any,
    *,
    run_document_id: str = "",
    run_document_name: str = "",
) -> str:
    """Pick storage_path from run-attached documents when file_path is unset."""
    normalized = normalize_run_documents(documents)
    ready: list[dict[str, Any]] = []
    for doc in normalized:
        if str(doc.get("status") or "") != "ready":
            continue
        path = str(doc.get("storage_path") or "").strip()
        if not path:
            continue
        if Path(path).suffix.lower() not in SPREADSHEET_STORAGE_SUFFIXES:
            continue
        ready.append(doc)
    if not ready:
        return ""
    doc_id = str(run_document_id or "").strip()
    if doc_id:
        for doc in ready:
            if str(doc.get("document_id") or "") == doc_id:
                return str(doc.get("storage_path") or "").strip()
        return ""
    name = str(run_document_name or "").strip()
    if name:
        for doc in ready:
            doc_name = str(doc.get("name") or "")
            if doc_name == name or doc_name.lower() == name.lower():
                return str(doc.get("storage_path") or "").strip()
        return ""
    if len(ready) == 1:
        return str(ready[0].get("storage_path") or "").strip()
    return ""


@dataclass(frozen=True)
class SpreadsheetRowRecord:
    row_number: int
    row_data: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "row_data": dict(self.row_data),
        }


@dataclass(frozen=True)
class SpreadsheetParseResult:
    source_file: str
    file_format: str
    sheet_name: str | None
    sheet_names: list[str]
    headers: list[str]
    rows: list[SpreadsheetRowRecord]

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def preview(self, *, limit: int = DEFAULT_SAMPLE_ROW_LIMIT) -> dict[str, Any]:
        return {
            "source_file": self.source_file,
            "file_format": self.file_format,
            "sheet_name": self.sheet_name,
            "sheet_names": list(self.sheet_names),
            "headers": list(self.headers),
            "row_count": self.row_count,
            "sample_rows": [row.to_dict() for row in self.rows[: max(0, int(limit))]],
        }


def infer_spreadsheet_format(file_path: str, file_format: str | None = None) -> str:
    normalized = str(file_format or "").strip().lower()
    if normalized == "auto":
        normalized = ""
    if normalized in SUPPORTED_SPREADSHEET_FORMATS:
        return normalized
    suffix = Path(file_path).suffix.lower().lstrip(".")
    if suffix in SUPPORTED_SPREADSHEET_FORMATS:
        return suffix
    raise SpreadsheetParseError("Spreadsheet format must be one of: csv, xlsx, or auto-detected from the file extension.")


def parse_spreadsheet(
    *,
    file_path: str,
    file_format: str | None = None,
    sheet_name: str | None = None,
    header_row_index: int = 1,
    start_row_index: int | None = None,
    empty_row_policy: str = "skip",
) -> SpreadsheetParseResult:
    normalized_path = str(file_path).strip()
    if not normalized_path:
        raise SpreadsheetParseError(
            "Spreadsheet file path is required. Set file_path on the Spreadsheet Rows node, or attach exactly one "
            "ready CSV/XLSX run document, or set run_document_id / run_document_name to choose among several."
        )
    path = Path(normalized_path).expanduser()
    if not path.exists() or not path.is_file():
        raise SpreadsheetParseError(f"Spreadsheet file not found: {normalized_path}")

    resolved_format = infer_spreadsheet_format(normalized_path, file_format)
    normalized_empty_policy = str(empty_row_policy or "skip").strip().lower()
    if normalized_empty_policy not in {"skip", "include"}:
        raise SpreadsheetParseError("Empty row policy must be either 'skip' or 'include'.")
    header_row_index = SPREADSHEET_HEADER_ROW_INDEX
    start_row_index = SPREADSHEET_FIRST_DATA_ROW_INDEX

    if resolved_format == "csv":
        return _parse_csv(
            path=path,
            header_row_index=header_row_index,
            start_row_index=start_row_index,
            empty_row_policy=normalized_empty_policy,
        )
    return _parse_xlsx(
        path=path,
        sheet_name=sheet_name,
        header_row_index=header_row_index,
        start_row_index=start_row_index,
        empty_row_policy=normalized_empty_policy,
    )


def _parse_csv(
    *,
    path: Path,
    header_row_index: int,
    start_row_index: int | None,
    empty_row_policy: str,
) -> SpreadsheetParseResult:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    return _build_parse_result(
        source_file=str(path),
        file_format="csv",
        sheet_name=None,
        sheet_names=[],
        raw_rows=rows,
        header_row_index=header_row_index,
        start_row_index=start_row_index,
        empty_row_policy=empty_row_policy,
    )


def _parse_xlsx(
    *,
    path: Path,
    sheet_name: str | None,
    header_row_index: int,
    start_row_index: int | None,
    empty_row_policy: str,
) -> SpreadsheetParseResult:
    if load_workbook is None:
        raise SpreadsheetParseError("XLSX support requires the 'openpyxl' package to be installed.")
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    selected_sheet_name = str(sheet_name or "").strip()
    if selected_sheet_name:
        if selected_sheet_name not in workbook.sheetnames:
            raise SpreadsheetParseError(
                f"Sheet '{selected_sheet_name}' was not found. Available sheets: {', '.join(sheet_names) or 'none'}."
            )
        worksheet = workbook[selected_sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]
        selected_sheet_name = worksheet.title
    raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return _build_parse_result(
        source_file=str(path),
        file_format="xlsx",
        sheet_name=selected_sheet_name,
        sheet_names=sheet_names,
        raw_rows=raw_rows,
        header_row_index=header_row_index,
        start_row_index=start_row_index,
        empty_row_policy=empty_row_policy,
    )


def _build_parse_result(
    *,
    source_file: str,
    file_format: str,
    sheet_name: str | None,
    sheet_names: list[str],
    raw_rows: Sequence[Sequence[Any]],
    header_row_index: int,
    start_row_index: int | None,
    empty_row_policy: str,
) -> SpreadsheetParseResult:
    if header_row_index > len(raw_rows):
        raise SpreadsheetParseError(
            f"Header row {header_row_index} is outside the available range of {len(raw_rows)} row(s)."
        )
    header_values = list(raw_rows[header_row_index - 1]) if raw_rows else []
    headers = _normalize_headers(header_values)
    first_data_row = start_row_index if start_row_index is not None else header_row_index + 1
    if first_data_row <= header_row_index:
        raise SpreadsheetParseError("First data row must come after the header row.")
    row_records: list[SpreadsheetRowRecord] = []
    for row_number in range(first_data_row, len(raw_rows) + 1):
        row_values = list(raw_rows[row_number - 1])
        normalized_values = [_normalize_cell_value(value) for value in row_values]
        row_data = {
            header: normalized_values[index] if index < len(normalized_values) else None
            for index, header in enumerate(headers)
        }
        if empty_row_policy == "skip" and _row_is_empty(row_data.values()):
            continue
        row_records.append(
            SpreadsheetRowRecord(
                row_number=row_number,
                row_data=row_data,
            )
        )
    return SpreadsheetParseResult(
        source_file=source_file,
        file_format=file_format,
        sheet_name=sheet_name,
        sheet_names=sheet_names,
        headers=headers,
        rows=row_records,
    )


def _normalize_headers(values: Sequence[Any]) -> list[str]:
    if not values:
        raise SpreadsheetParseError("Header row is empty.")
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        candidate = str(value or "").strip()
        if not candidate:
            candidate = f"column_{index}"
        candidate = candidate.replace("\n", " ").replace("\r", " ").strip()
        candidate = "_".join(part for part in candidate.split(" ") if part)
        normalized = candidate or f"column_{index}"
        count = seen.get(normalized, 0) + 1
        seen[normalized] = count
        headers.append(normalized if count == 1 else f"{normalized}_{count}")
    return headers


def _normalize_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return value


def _row_is_empty(values: Iterable[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True
