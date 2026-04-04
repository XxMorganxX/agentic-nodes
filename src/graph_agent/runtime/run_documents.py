from __future__ import annotations

import csv
import io
import json
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping
from uuid import uuid4

DEFAULT_RUN_DOCUMENT_UPLOAD_DIR = Path(__file__).resolve().parents[3] / ".graph-agent" / "uploads"
DOCUMENT_EXCERPT_LIMIT = 1200
SUPPORTED_DOCUMENT_EXTENSIONS = {".csv", ".json", ".md", ".markdown", ".pdf", ".txt", ".xlsx"}
SAFE_FILENAME_PATTERN = re.compile(r"[^A-Za-z0-9._-]+")


class RunDocumentIngestionError(ValueError):
    pass


@dataclass
class RunDocumentRecord:
    document_id: str
    name: str
    mime_type: str
    size_bytes: int
    storage_path: str
    text_content: str
    text_excerpt: str
    status: str
    error: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "document_id": self.document_id,
            "name": self.name,
            "mime_type": self.mime_type,
            "size_bytes": self.size_bytes,
            "storage_path": self.storage_path,
            "text_content": self.text_content,
            "text_excerpt": self.text_excerpt,
            "status": self.status,
            "error": self.error,
        }


def resolve_run_document_upload_dir() -> Path:
    configured = os.environ.get("GRAPH_AGENT_UPLOAD_DIR", "").strip()
    if configured:
        return Path(configured).expanduser()
    return DEFAULT_RUN_DOCUMENT_UPLOAD_DIR


def normalize_run_documents(documents: Any) -> list[dict[str, Any]]:
    if not isinstance(documents, Iterable) or isinstance(documents, (str, bytes, bytearray, Mapping)):
        return []
    normalized: list[dict[str, Any]] = []
    for candidate in documents:
        if not isinstance(candidate, Mapping):
            continue
        document_id = str(candidate.get("document_id") or "").strip()
        name = str(candidate.get("name") or "").strip()
        if not document_id or not name:
            continue
        normalized.append(
            {
                "document_id": document_id,
                "name": name,
                "mime_type": str(candidate.get("mime_type") or "application/octet-stream"),
                "size_bytes": _coerce_int(candidate.get("size_bytes")),
                "storage_path": str(candidate.get("storage_path") or ""),
                "text_content": str(candidate.get("text_content") or ""),
                "text_excerpt": str(candidate.get("text_excerpt") or ""),
                "status": str(candidate.get("status") or "ready"),
                "error": _coerce_optional_string(candidate.get("error")),
            }
        )
    return normalized


def ingest_run_documents(documents: Iterable[Mapping[str, Any]], *, storage_root: Path | None = None) -> list[dict[str, Any]]:
    root = storage_root or resolve_run_document_upload_dir()
    root.mkdir(parents=True, exist_ok=True)
    ingested: list[dict[str, Any]] = []
    for candidate in documents:
        ingested.append(_ingest_single_document(candidate, root).to_dict())
    return ingested


def _ingest_single_document(document: Mapping[str, Any], storage_root: Path) -> RunDocumentRecord:
    name = str(document.get("name") or "").strip()
    if not name:
        raise RunDocumentIngestionError("Document name is required.")
    raw_bytes = document.get("data")
    if not isinstance(raw_bytes, (bytes, bytearray)):
        raise RunDocumentIngestionError(f"Document '{name}' is missing file bytes.")
    content_type = str(document.get("content_type") or "").strip() or "application/octet-stream"
    document_id = uuid4().hex
    safe_name = _safe_filename(name)
    raw_bytes = bytes(raw_bytes)
    extension = Path(name).suffix.strip().lower()

    if extension == ".xlsx":
        stored_name = f"{Path(name).stem or 'spreadsheet'}.csv"
        try:
            workbook = _load_xlsx_workbook(raw_bytes)
            csv_bytes = _xlsx_first_sheet_to_csv_bytes(workbook)
            text_content = _xlsx_workbook_to_prompt_text(workbook)
        except RunDocumentIngestionError as exc:
            excerpt = _excerpt_for_document("", str(exc))
            return RunDocumentRecord(
                document_id=document_id,
                name=name,
                mime_type=content_type,
                size_bytes=len(raw_bytes),
                storage_path="",
                text_content="",
                text_excerpt=excerpt,
                status="failed",
                error=str(exc),
            )
        csv_safe = _safe_filename(stored_name)
        storage_path = storage_root / f"{document_id}-{csv_safe}"
        storage_path.write_bytes(csv_bytes)
        excerpt = _excerpt_for_document(text_content, None)
        return RunDocumentRecord(
            document_id=document_id,
            name=stored_name,
            mime_type="text/csv",
            size_bytes=len(csv_bytes),
            storage_path=str(storage_path),
            text_content=text_content,
            text_excerpt=excerpt,
            status="ready",
            error=None,
        )

    storage_path = storage_root / f"{document_id}-{safe_name}"
    storage_path.write_bytes(raw_bytes)
    try:
        text_content = _extract_document_text(name, raw_bytes)
        status = "ready"
        error = None
    except RunDocumentIngestionError as exc:
        text_content = ""
        status = "failed"
        error = str(exc)
    excerpt = _excerpt_for_document(text_content, error)
    return RunDocumentRecord(
        document_id=document_id,
        name=name,
        mime_type=content_type,
        size_bytes=len(raw_bytes),
        storage_path=str(storage_path),
        text_content=text_content,
        text_excerpt=excerpt,
        status=status,
        error=error,
    )


def _extract_document_text(name: str, raw_bytes: bytes) -> str:
    extension = Path(name).suffix.strip().lower()
    if extension not in SUPPORTED_DOCUMENT_EXTENSIONS:
        raise RunDocumentIngestionError(
            "Unsupported document type. Upload .txt, .md, .markdown, .json, .csv, .xlsx, or .pdf files."
        )
    if extension in {".txt", ".md", ".markdown"}:
        return raw_bytes.decode("utf-8-sig", errors="replace").strip()
    if extension == ".json":
        return _normalize_json_text(raw_bytes)
    if extension == ".csv":
        return _normalize_csv_text(raw_bytes)
    if extension == ".xlsx":
        return _extract_xlsx_text(raw_bytes)
    if extension == ".pdf":
        return _extract_pdf_text(raw_bytes)
    raise RunDocumentIngestionError(f"Unsupported document type '{extension}'.")


def _normalize_json_text(raw_bytes: bytes) -> str:
    decoded = raw_bytes.decode("utf-8-sig", errors="replace")
    try:
        payload = json.loads(decoded)
    except json.JSONDecodeError:
        return decoded.strip()
    return json.dumps(payload, indent=2, sort_keys=True, ensure_ascii=True)


def _normalize_csv_text(raw_bytes: bytes) -> str:
    decoded = raw_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(decoded))
    rows = [[cell.strip() for cell in row] for row in reader]
    rendered_rows = [", ".join(cell for cell in row if cell) for row in rows if any(cell for cell in row)]
    return "\n".join(rendered_rows).strip()


def _load_xlsx_workbook(raw_bytes: bytes):
    try:
        import openpyxl
    except ImportError as exc:
        raise RunDocumentIngestionError("XLSX support requires the 'openpyxl' package.") from exc
    try:
        return openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise RunDocumentIngestionError(f"Unable to read XLSX file: {exc}") from exc


def _xlsx_workbook_to_prompt_text(workbook) -> str:
    sheet_texts: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(cells):
                rows.append(", ".join(cells))
        if rows:
            sheet_texts.append(f"Sheet: {sheet.title}\n" + "\n".join(rows))
    rendered = "\n\n".join(sheet_texts).strip()
    if not rendered:
        raise RunDocumentIngestionError("Uploaded XLSX file contained no readable data.")
    return rendered


def _xlsx_first_sheet_to_csv_bytes(workbook) -> bytes:
    worksheets = workbook.worksheets
    if not worksheets:
        raise RunDocumentIngestionError("Uploaded XLSX file has no worksheets.")
    for sheet in worksheets:
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        row_count = 0
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if cell is None else str(cell) for cell in row])
            row_count += 1
        if row_count > 0:
            return buffer.getvalue().encode("utf-8-sig")
    raise RunDocumentIngestionError("Uploaded XLSX contained no rows to export as CSV.")


def _extract_xlsx_text(raw_bytes: bytes) -> str:
    workbook = _load_xlsx_workbook(raw_bytes)
    return _xlsx_workbook_to_prompt_text(workbook)


def _extract_pdf_text(raw_bytes: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - dependency verified in integration paths
        raise RunDocumentIngestionError("PDF support requires the 'pypdf' package.") from exc
    try:
        reader = PdfReader(io.BytesIO(raw_bytes))
        page_text = [page.extract_text() or "" for page in reader.pages]
    except Exception as exc:  # noqa: BLE001
        raise RunDocumentIngestionError(f"Unable to extract text from PDF: {exc}") from exc
    rendered = "\n\n".join(text.strip() for text in page_text if text and text.strip())
    if not rendered:
        raise RunDocumentIngestionError("Uploaded PDF did not contain readable text.")
    return rendered


def _excerpt_for_document(text_content: str, error: str | None) -> str:
    if text_content.strip():
        excerpt = text_content.strip()
        if len(excerpt) <= DOCUMENT_EXCERPT_LIMIT:
            return excerpt
        return f"{excerpt[: DOCUMENT_EXCERPT_LIMIT - 3].rstrip()}..."
    return error or ""


def _safe_filename(name: str) -> str:
    base_name = Path(name).name.strip() or "document"
    sanitized = SAFE_FILENAME_PATTERN.sub("-", base_name).strip("-.")
    return sanitized or "document"


def _coerce_int(value: Any) -> int:
    try:
        return max(int(value), 0)
    except (TypeError, ValueError):
        return 0


def _coerce_optional_string(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
